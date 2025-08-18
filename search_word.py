import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

# MAX_PATH = 240  # Windowsのパス長制限


def is_include_search_word(dir_path, file_name, search_results):
    nec_word = include_nec_search_word(f"{dir_path}/{file_name}")
    if nec_word:
        # NEC関連の検索ワードが含まれている場合は、結果に追加
        search_results.append(
            {
                "SearchWord": nec_word,
                "Directory": dir_path,
                "FileName": file_name,
            },
        )
        return True

    word = include_search_word(file_name)
    if word:
        # 検索ワードが含まれている場合は、結果に追加
        search_results.append(
            {
                "SearchWord": word,
                "Directory": dir_path,
                "FileName": file_name,
            },
        )
        return True

    return False


# def get_safe_save_path(output_dir, file_path):
#     # サブディレクトリ構造を維持しつつ、パス長が長すぎる場合はファイル名をハッシュ化
#     # save_path = Path(output_dir) / file_path.lstrip("/")
#     save_path = output_dir / file_path.lstrip("/")
#     hash_name = ""
#     if len(str(save_path)) > MAX_PATH:
#         p = Path(file_path)
#         ext = p.suffix
#         hash_name = hashlib.sha256(file_path.encode("utf-8")).hexdigest() + ext
#         save_path = Path(output_dir) / hash_name
#     return save_path, hash_name


def read_search_words(file_path):
    """検索ワードを外部ファイルから読み込む"""
    if not Path(file_path).exists():
        print(f"検索ワードファイルが見つかりません: {file_path}")
        sys.exit(1)
    with open(file_path, encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]


def read_exclusion_words(file_path):
    """除外ワードを外部ファイルから読み込む"""
    if not Path(file_path).exists():
        print(f"除外ワードファイルが見つかりません: {file_path}")
        sys.exit(1)
    with open(file_path, encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]


def generate_regex(word):
    """単語から大文字小文字・全角半角を許容する正規表現を生成する"""
    regex_parts = [
        f"[{char.lower()}{char.upper()}{chr(ord(char.lower()) + 0xFEE0)}{chr(ord(char.upper()) + 0xFEE0)}]"
        for char in word
        if char.isalpha()
    ]
    return r"(?:{})".format("".join(regex_parts))


search_words_file = "search_words.txt"
search_words = read_search_words(search_words_file)
search_nec_words_file = "search_nec_words.txt"
search_nec_words = read_search_words(search_nec_words_file)
merged_words = search_nec_words + search_words
exclusion_words_file = "exclusion_words.txt"
exclusion_words = read_exclusion_words(exclusion_words_file)


def include_search_word(filename):
    """ファイル名に検索ワードが含まれているかどうかをチェックする関数"""
    results = False

    for word in search_words:
        if re.search(word, filename, re.IGNORECASE):
            results = word
            break

    return results


def include_nec_search_word(line):
    """NEC関連の検索ワードが含まれているかどうかをチェックする関数"""
    result = False

    nec_match = False
    for word in search_nec_words:
        # NEC社名関連ワード
        if word == "NEC":
            matches = list(re.finditer(generate_regex(word), line))
            if matches:
                # 各マッチ部分について、除外ワードを含まないものが1つでもあればヒット
                for m in matches:
                    matched_text = m.group()
                    if not any(re.search(generate_regex(ex_word), matched_text) for ex_word in exclusion_words):
                        result = word
                        break
                else:
                    # 全てのマッチが除外ワードを含む場合は無視
                    return False
                break
        # 間に0文字以上を許容し、すべての文字に対して大文字小文字・全角半角を許容
        elif word in "nippondenki":
            nec_match = re.search(
                r"(?:[nNｎＮ][iIｉＩ][pPｐＰ][pPｐＰ][oOｏＯ][nNｎＮ].*?[dDｄＤ][eEｅＥ][nNｎＮ][kKｋＫ][iIｉＩ])",
                line,
            )
            if nec_match:
                result = word
                break
        # 間に0文字以上を許容し、すべての文字に対して大文字小文字・全角半角を許容
        elif word in "nipponelectric":
            nec_match = re.search(
                r"(?:[nNｎＮ][iIｉＩ][pPｐＰ][pPｐＰ][oOｏＯ][nNｎＮ].*?[eEｅＥ][lLｌＬ][eEｅＥ][cCｃＣ][tTｔＴ][rRｒＲ][iIｉＩ][cCｃＣ])",
                line,
            )
            if nec_match:
                result = word
                break
        # 間に0文字以上を許容
        elif word in "日本電気":
            nec_match = re.search(r"(?:日本.*?電気)", line)
            if nec_match:
                result = word
                break
        # 間に0文字以上を許容
        elif word in "にほんでんき":
            nec_match = re.search(
                r"(?:にほん.*?でんき)",
                line,
            )
            if nec_match:
                result = word
                break
        # 間に0文字以上を許容
        elif word in "にっぽんでんき":
            nec_match = re.search(
                r"(?:にっぽん.*?でんき)",
                line,
            )
            if nec_match:
                result = word
                break
        # 間に0文字以上を許容
        elif word in "ニホンデンキ":
            nec_match = re.search(
                r"(?:[ﾆニ][ﾎホ][ﾝン].*?[ﾃﾞデ][ﾝン][ｷキ])",
                line,
            )
            if nec_match:
                result = word
                break
        # 間に0文字以上を許容
        elif word in "ニッポンデンキ":
            nec_match = re.search(
                r"(?:[ﾆニ][ｯッ][ﾎホ][ﾝン].*?[ﾃﾞデ][ﾝン][ｷキ])",
                line,
            )
            if nec_match:
                result = word
                break
        elif word in "日電":
            nec_match = re.search(r"(?:日電)", line)
            if nec_match:
                result = word
                break
        elif word in "にちでん":
            nec_match = re.search(r"(?:にちでん)", line)
            if nec_match:
                result = word
                break
        elif word in "ニチデン":
            nec_match = re.search(
                r"(?:[ﾆニ][ﾁチ][ﾃﾞデ][ﾝン])",
                line,
            )
            if nec_match:
                result = word
                break

    return result


def write_to_xlsx(results, output_file_path):
    """結果をXLSXファイルに書き込む"""
    wb = openpyxl.Workbook()

    # 1シート目: 検索結果
    ws1 = wb.active
    ws1.title = "Search Results"

    # ヘッダーの書き込み
    headers = ["SearchWord", "Directory", "FileName", "HashName"]
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # データの書き込み
    for row_num, result in enumerate(results, 2):
        ws1.cell(row=row_num, column=1, value=result["SearchWord"])
        ws1.cell(row=row_num, column=2, value=result["Directory"])
        ws1.cell(row=row_num, column=3, value=result["FileName"])
        # ws1.cell(row=row_num, column=4, value=result["HashName"])

    # 2シート目: 検索ワードごとのカウント
    ws2 = wb.create_sheet(title="Search Word Counts")

    # カウントの計算
    word_counts = dict.fromkeys(merged_words, 0)
    for result in results:
        word_counts[result["SearchWord"]] += 1

    # ヘッダーの書き込み
    ws2.cell(row=1, column=1, value="SearchWord").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Count").font = Font(bold=True)

    # データの書き込み
    for row_num, (word, count) in enumerate(word_counts.items(), 2):
        ws2.cell(row=row_num, column=1, value=word)
        ws2.cell(row=row_num, column=2, value=count)

    # ファイルを保存
    wb.save(output_file_path)
    print(f"結果をXLSX形式で出力しました: {output_file_path}")


def keyword_index(result):
    try:
        return merged_words.index(result["SearchWord"])
    except ValueError:
        return len(merged_words)
